import openpyxl
from openpyxl.styles import Color, Style, PatternFill


def open_py_excel_pass(index, ws_index):
    wb = openpyxl.load_workbook("C:\Users\sagarkul\Desktop\DataSource.xlsx")
    ws = wb.worksheets[ws_index]
    ws.cell(row=index + 1, column=2).value = "Pass"
    ws.cell(row=index + 1, column=2).style = Style(fill=PatternFill(patternType='solid', fgColor=Color('00cc00')))
    wb.save("C:\Users\sagarkul\Desktop\DataSource.xlsx")


def open_py_excel_fail(index, ws_index):
    wb = openpyxl.load_workbook("C:\Users\sagarkul\Desktop\DataSource.xlsx")
    ws = wb.worksheets[ws_index]
    ws.cell(row=index + 1, column=2).value = "Fail"
    ws.cell(row=index + 1, column=2).style = Style(fill=PatternFill(patternType='solid', fgColor=Color('ff1a1a')))
    wb.save("C:\Users\sagarkul\Desktop\DataSource.xlsx")
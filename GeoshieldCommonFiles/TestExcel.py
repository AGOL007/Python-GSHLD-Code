
import openpyxl
from openpyxl.styles import Color, Alignment, Style, PatternFill


def open_py_excel_pass(sheet_index, index):
    wb = openpyxl.load_workbook("C:\Users\sagarkul\Desktop\TestCase.xlsx")
    ws = wb.worksheets[sheet_index]
    for i in range(2,10):
        if ws.cell(row=i, column=index+1).value == "Pass":
            continue
        elif ws.cell(row=i, column=index+1).value == "Fail":
            continue
        else:
            ws.cell(row=i, column=index+1).value = "Pass"
            ws.cell(row=i, column=index+1).style = Style(fill=PatternFill(patternType='solid', fgColor=Color('009900')))
            ws.cell(row=1, column=index+1).value = index
            break
    wb.save("C:\Users\sagarkul\Desktop\TestCase.xlsx")


def open_py_excel_fail():
    wb = openpyxl.load_workbook("C:\Users\sagarkul\Desktop\TestCase.xlsx")
    ws = wb.worksheets[0]
    for i in range(2,10):
        if ws.cell(row=i, column=2).value == "Pass":
           continue
        if ws.cell(row=i, column=2).value == "Fail":
           continue
        else:
            ws.cell(row=i, column=2).value = "Fail"
            break
    wb.save("C:\Users\sagarkul\Desktop\TestCase.xlsx")
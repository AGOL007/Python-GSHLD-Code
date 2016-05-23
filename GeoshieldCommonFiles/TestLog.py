# -*- coding: utf-8 -*-
""" Unit test file """
__author__ = 'SagarKul'

import os.path
import time
import logging
import xlwt
from openpyxl import Workbook
import TestLogBombBlast


def create_results_bomb_item(tool_bomb_blast_bomb_type, tool_name):
    if os.path.exists(r"C:\Users\sagarkul\Desktop\Test_results.xls"):
        file = open(r"C:\Users\sagarkul\Desktop\Test_results.xls", "a")
    else:
        file = open(r"C:\Users\sagarkul\Desktop\Test_results.xls", "w")
    show_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    file.write("........Test_Result........" + tool_name + show_time)
    file.write("\n")
    file.write("bomb type " + tool_bomb_blast_bomb_type + " is selected\n")
    file.close()


def create_result_validation_right_panel_layer(layer_name):
    if os.path.exists(r"C:\Users\sagarkul\Desktop\Test_results.xls"):
        file = open(r"C:\Users\sagarkul\Desktop\Test_results.xls", "a")
    else:
        file = open(r"C:\Users\sagarkul\Desktop\Test_results.xls", "w")
    file.write(layer_name + " right panel layer gets validated\n")
    file.write("\n")


def create_results_validation_panel_feature_count():
    if os.path.exists("C:\Users\sagarkul\Desktop\Test_results.xls"):
        file = open("C:\Users\sagarkul\Desktop\Test_results.xls", "a")
    else:
        file = open("C:\Users\sagarkul\Desktop\Test_results.xls", "w")
    file.write(" panel feature count gets validated\n")
    file.write("\n")


def create_results_validation_fature_on_map(layer_name):
    if os.path.exists("C:\Users\sagarkul\Desktop\Test_results.xls"):
        file = open("C:\Users\sagarkul\Desktop\Test_results.xls", "a")
    else:
        file = open("C:\Users\sagarkul\Desktop\Test_results.xls", "w")
    file.write(layer_name + " feature present on map\n")
    file.write("\n")


def write_test_case_result(result, location):
    w = xlwt.Workbook()
    wb = Workbook()
    ws1 = wb.worksheets[0]
    ws1.title = 'Sheet1'
    dest_filename = r'C:\Users\sagarkul\Desktop'

    while True:
        if result == "PASS":
            ws1.cell(location).value = "PASSED"
            break
        else:
            ws1.cell(location).value = "FAILED"
            break
    # Save the file
    wb.save(filename = dest_filename)


def log_creation(tool_name, method_name, index, ws_index):
    operation = 'w'
    if os.path.exists(r"C:\Users\sagarkul\Desktop\Test_log.txt"):
        operation = 'a'
    with open(r"C:\Users\sagarkul\Desktop\Test_log.txt", operation) as file_:
        file_.write("-"*50)
        file_.write("\n")
        show_time = time.strftime("%Y-%m-%d %I:%M:%S", time.localtime())
        file_.write("........Test_log........ " + tool_name + "  " + show_time)
        file_.write("\n")
        file_.write("For DataRow index " + str(index) + " ")
        file_.write("\n")
        file_.write(method_name + " " + "Method gets failed.")
        file_.write("\n")
    #    file_.write(val)
    #    file.write("......................................................")
        file_.write("\n")
        TestLogBombBlast.open_py_excel_fail(index, ws_index)
        assert False


def create_log_using_module(index):
    logging.basicConfig(filename=r"C:\Users\sagarkul\Desktop\sample.txt", level=logging.ERROR)
    logging.info("For DataRow index {}".format(index))

